"""
IKEA SERVER HEALTH CHECK MONITORING TOOL - UNIFIED RUNNER
==========================================================

Single script to run all health checks, generate reports, and create Excel/HTML outputs.
No external dependencies required - everything is self-contained.

Usage:
    python run.py              # Run single check cycle
    python run.py --continuous # Run continuous monitoring
    python run.py --report     # Generate report from latest data
"""

import requests
import time
import logging
import json
import csv
import sys
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Tuple
from collections import defaultdict
import configparser

# ============================================================================
# CONFIGURATION LOADER
# ============================================================================

class ConfigLoader:
    """Load configuration from config.ini"""
    
    @staticmethod
    def load():
        """Load and return configuration dictionary"""
        config = configparser.ConfigParser()
        config_file = Path("config.ini")
        
        if not config_file.exists():
            return ConfigLoader.get_defaults()
        
        config.read(config_file)
        
        return {
            'check_interval': config.getint('MONITORING', 'check_interval', fallback=300),
            'quick_check_interval': config.getint('MONITORING', 'quick_check_interval', fallback=60),
            'request_timeout': config.getint('MONITORING', 'request_timeout', fallback=10),
            'alert_threshold': config.getint('MONITORING', 'alert_threshold', fallback=2),
            'log_directory': config.get('LOGGING', 'log_directory', fallback='logs'),
            'log_file': config.get('LOGGING', 'log_file', fallback='health_check.log'),
            'alert_log_file': config.get('LOGGING', 'alert_log_file', fallback='health_check_alerts.log'),
            'report_file': config.get('LOGGING', 'report_file', fallback='health_check_report.json'),
            'email_enabled': config.getboolean('EMAIL_ALERTS', 'enabled', fallback=False),
            'use_outlook': config.getboolean('EMAIL_ALERTS', 'use_outlook', fallback=True),
            'use_smtp': config.getboolean('EMAIL_ALERTS', 'use_smtp', fallback=False),
            'sender_email': config.get('EMAIL_ALERTS', 'sender_email', fallback=''),
            'recipient_emails': [e.strip() for e in config.get('EMAIL_ALERTS', 'recipient_emails', fallback='').split(',')],
            'smtp_server': config.get('EMAIL_ALERTS', 'smtp_server', fallback='smtp.gmail.com'),
            'smtp_port': config.getint('EMAIL_ALERTS', 'smtp_port', fallback=587),
            'healthy_status_codes': [int(c.strip()) for c in config.get('HTTP_STATUS', 'healthy_status_codes', fallback='200,201,202,204,301,302,304,307,308').split(',')]
        }
    
    @staticmethod
    def get_defaults():
        """Return default configuration"""
        return {
            'check_interval': 300,
            'quick_check_interval': 60,
            'request_timeout': 10,
            'alert_threshold': 2,
            'log_directory': 'logs',
            'log_file': 'health_check.log',
            'alert_log_file': 'health_check_alerts.log',
            'report_file': 'health_check_report.json',
            'email_enabled': False,
            'use_outlook': True,
            'use_smtp': False,
            'sender_email': '',
            'recipient_emails': [],
            'smtp_server': 'smtp.gmail.com',
            'smtp_port': 587,
            'healthy_status_codes': [200, 201, 202, 204, 301, 302, 304, 307, 308]
        }


# ============================================================================
# LOGGER SETUP
# ============================================================================

def setup_logging(config: Dict):
    """Configure logging for both console and file output"""
    log_dir = Path(config['log_directory'])
    log_dir.mkdir(exist_ok=True)
    
    # Main logger
    logger = logging.getLogger("HealthCheck")
    logger.setLevel(logging.DEBUG)
    logger.handlers.clear()  # Clear any existing handlers
    
    # File handler
    file_handler = logging.FileHandler(log_dir / config['log_file'])
    file_handler.setLevel(logging.DEBUG)
    
    # Console handler
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    
    # Formatter
    formatter = logging.Formatter(
        '%(asctime)s - %(levelname)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    file_handler.setFormatter(formatter)
    console_handler.setFormatter(formatter)
    
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    
    # Alert logger
    alert_logger = logging.getLogger("AlertLog")
    alert_logger.setLevel(logging.WARNING)
    alert_logger.handlers.clear()
    alert_handler = logging.FileHandler(log_dir / config['alert_log_file'])
    alert_handler.setFormatter(formatter)
    alert_logger.addHandler(alert_handler)
    
    return logger, alert_logger


# ============================================================================
# HEALTH CHECK MONITOR
# ============================================================================

class HealthCheckMonitor:
    """Main health check monitoring engine"""
    
    def __init__(self, config: Dict):
        """Initialize the health check monitor"""
        self.config = config
        self.logger, self.alert_logger = setup_logging(config)
        self.urls = self._load_urls()
        self.logger.info(f"Loaded {len(self.urls)} URLs for monitoring")
        
        # Status tracking
        self.status_history: Dict[str, List[Dict]] = defaultdict(list)
        self.consecutive_failures: Dict[str, int] = defaultdict(int)
        self.downtime_start: Dict[str, datetime] = {}
        
        # Statistics
        self.total_checks = 0
        self.total_failures = 0
        self.start_time = datetime.now()
    
    def _load_urls(self) -> List[str]:
        """Load URLs from file"""
        try:
            with open("urls.txt", 'r') as f:
                urls = [line.strip() for line in f if line.strip()]
            return urls
        except FileNotFoundError:
            self.logger.error("urls.txt not found")
            return []
    
    def check_url_health(self, url: str) -> Tuple[bool, int, float, str]:
        """Check the health of a single URL"""
        try:
            start_time = time.time()
            response = requests.get(
                url,
                timeout=self.config['request_timeout'],
                allow_redirects=True,
                verify=False  # Ignore SSL certificate warnings
            )
            response_time = time.time() - start_time
            
            is_healthy = response.status_code in self.config['healthy_status_codes']
            return is_healthy, response.status_code, response_time, "OK"
        
        except requests.exceptions.Timeout:
            return False, 0, self.config['request_timeout'], "Timeout"
        except requests.exceptions.ConnectionError:
            return False, 0, 0, "Connection Error"
        except requests.exceptions.RequestException as e:
            return False, 0, 0, f"Request Error"
        except Exception as e:
            return False, 0, 0, f"Error"
    
    def record_status(self, url: str, is_healthy: bool, status_code: int, 
                     response_time: float, message: str):
        """Record the status check result"""
        status_entry = {
            "timestamp": datetime.now().isoformat(),
            "url": url,
            "status": "UP" if is_healthy else "DOWN",
            "status_code": status_code,
            "response_time": f"{response_time:.3f}s",
            "message": message
        }
        
        self.status_history[url].append(status_entry)
        
        # Keep only last 1000 entries per URL
        if len(self.status_history[url]) > 1000:
            self.status_history[url] = self.status_history[url][-1000:]
    
    def handle_failure(self, url: str, status_code: int, message: str):
        """Handle URL health check failure"""
        self.consecutive_failures[url] += 1
        
        self.logger.warning(f"FAILED: {url} - Code: {status_code} - {message}")
        
        if self.consecutive_failures[url] == 1:
            self.downtime_start[url] = datetime.now()
        
        if self.consecutive_failures[url] == self.config['alert_threshold']:
            self._send_alert(url, status_code, message)
    
    def handle_success(self, url: str):
        """Handle URL health check success"""
        was_down = self.consecutive_failures[url] >= self.config['alert_threshold']
        
        if was_down:
            downtime_duration = datetime.now() - self.downtime_start.get(url, datetime.now())
            self.logger.info(f"[RECOVERY] {url} UP after {downtime_duration.total_seconds():.0f}s downtime")
            self._send_recovery_alert(url, downtime_duration)
        
        self.consecutive_failures[url] = 0
        self.downtime_start.pop(url, None)
    
    def _send_alert(self, url: str, status_code: int, message: str):
        """Send alert for URL failure"""
        alert_msg = f"[ALERT] {url} DOWN - Code: {status_code} - {message}"
        self.alert_logger.error(alert_msg)
        self.logger.error(alert_msg)
    
    def _send_recovery_alert(self, url: str, downtime_duration: timedelta):
        """Send alert for URL recovery"""
        alert_msg = f"[RECOVERY] {url} UP after {downtime_duration.total_seconds():.0f}s downtime"
        self.alert_logger.info(alert_msg)
        self.logger.warning(alert_msg)
    
    def run_single_check_cycle(self):
        """Run a single health check cycle for all URLs"""
        print("\n" + "=" * 100)
        print(f"  HEALTH CHECK CYCLE - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print("=" * 100)
        
        for url in self.urls:
            is_healthy, status_code, response_time, message = self.check_url_health(url)
            
            self.record_status(url, is_healthy, status_code, response_time, message)
            self.total_checks += 1
            
            if is_healthy:
                status_icon = "[OK]"
                print(f"  {status_icon:6} | {url:45} | Code: {status_code} | Time: {response_time:.3f}s")
                self.handle_success(url)
            else:
                status_icon = "[FAIL]"
                print(f"  {status_icon:6} | {url:45} | Error: {message}")
                self.handle_failure(url, status_code, message)
                self.total_failures += 1
        
        # Generate report
        self._generate_report()
        print("=" * 100)
    
    def _generate_report(self):
        """Generate and save health check report"""
        report_data = {
            "timestamp": datetime.now().isoformat(),
            "monitoring_duration_seconds": (datetime.now() - self.start_time).total_seconds(),
            "total_checks": self.total_checks,
            "total_failures": self.total_failures,
            "failure_rate": f"{(self.total_failures / max(self.total_checks, 1) * 100):.2f}%",
            "monitored_urls": len(self.urls),
            "url_status_summary": {}
        }
        
        for url in self.urls:
            latest_status = self.status_history[url][-1] if self.status_history[url] else {}
            failed_count = self.consecutive_failures[url]
            
            report_data["url_status_summary"][url] = {
                "current_status": latest_status.get("status", "UNKNOWN"),
                "consecutive_failures": failed_count,
                "last_check": latest_status.get("timestamp", "N/A"),
                "last_status_code": latest_status.get("status_code", "N/A"),
                "last_response_time": latest_status.get("response_time", "N/A")
            }
        
        # Save report
        log_dir = Path(self.config['log_directory'])
        report_path = log_dir / self.config['report_file']
        with open(report_path, 'w') as f:
            json.dump(report_data, f, indent=2)
        
        return report_data
    
    def get_status_summary(self) -> Dict:
        """Get current status summary"""
        summary = {
            "timestamp": datetime.now().isoformat(),
            "healthy_urls": 0,
            "unhealthy_urls": 0,
            "urls_status": {}
        }
        
        for url in self.urls:
            latest_status = self.status_history[url][-1] if self.status_history[url] else None
            
            if latest_status:
                is_healthy = latest_status["status"] == "UP"
                if is_healthy:
                    summary["healthy_urls"] += 1
                else:
                    summary["unhealthy_urls"] += 1
                
                summary["urls_status"][url] = latest_status
        
        return summary
    
    def print_summary(self):
        """Print and log monitoring summary only once at the end"""
        duration = datetime.now() - self.start_time
        healthy = sum(1 for url in self.urls if self.consecutive_failures[url] == 0)
        down = len(self.urls) - healthy
        summary_lines = [
            "\n" + "=" * 100,
            "  MONITORING SUMMARY",
            "=" * 100,
            f"  Duration:           {duration}",
            f"  Total Checks:       {self.total_checks}",
            f"  Total Failures:     {self.total_failures}",
            f"  Failure Rate:       {(self.total_failures / max(self.total_checks, 1) * 100):.2f}%",
            f"  URLs Monitored:     {len(self.urls)}",
            f"  Currently Healthy:  {healthy}",
            f"  Currently Down:     {down}",
            "=" * 100 + "\n"
        ]
        summary_text = "\n".join(summary_lines)
        print(summary_text)
        self.logger.info(summary_text)
    
    def start_continuous_monitoring(self, duration_hours: int = None):
        """Start continuous monitoring"""
        print("\n" + "=" * 100)
        print("  STARTING CONTINUOUS MONITORING")
        print("=" * 100)
        if duration_hours:
            print(f"  Will run for {duration_hours} hours")
        else:
            print("  Running indefinitely (press Ctrl+C to stop)")
        print("=" * 100 + "\n")
        
        start_time = datetime.now()
        end_time = start_time + timedelta(hours=duration_hours) if duration_hours else None
        
        try:
            while True:
                self.run_single_check_cycle()
                
                if end_time and datetime.now() >= end_time:
                    break
                
                # Determine wait time
                wait_time = self.config['check_interval']
                if any(self.consecutive_failures[url] > 0 for url in self.urls):
                    wait_time = self.config['quick_check_interval']
                
                print(f"\n  Next check in {wait_time} seconds...\n")
                time.sleep(wait_time)
        
        except KeyboardInterrupt:
            print("\n\n  Monitoring stopped by user\n")
        except Exception as e:
            self.logger.error(f"Error during monitoring: {str(e)}")
        finally:
            self.print_summary()


# ============================================================================
# REPORT GENERATOR
# ============================================================================

class ReportGenerator:
    """Generate Excel and HTML reports"""
    
    def __init__(self, config: Dict):
        self.config = config
        self.log_dir = Path(config['log_directory'])
        self.report_file = self.log_dir / config['report_file']
    
    def generate_excel_report(self):
        """Generate Excel report from latest data"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            print("  ⚠ openpyxl not installed. Skipping Excel report.")
            return
        
        if not self.report_file.exists():
            print("  ⚠ No report data available yet. Run a health check first.")
            return
        
        with open(self.report_file, 'r') as f:
            data = json.load(f)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Health Check Report"
        
        # Header styling
        header_fill = PatternFill(start_color="0051BA", end_color="0051BA", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        center_align = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Summary section
        ws['A1'] = "IKEA HEALTH CHECK REPORT"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')
        
        ws['A3'] = f"Report Generated: {data['timestamp']}"
        ws['A4'] = f"Monitoring Duration: {data['monitoring_duration_seconds']:.0f} seconds"
        ws['A5'] = f"Total Checks: {data['total_checks']}"
        ws['A6'] = f"Total Failures: {data['total_failures']}"
        ws['A7'] = f"Failure Rate: {data['failure_rate']}"
        
        # URL Status Table Header
        ws['A9'] = "URL"
        ws['B9'] = "Status"
        ws['C9'] = "Response Time"
        ws['D9'] = "Consecutive Failures"
        ws['E9'] = "Last Check"
        
        for cell in ['A9', 'B9', 'C9', 'D9', 'E9']:
            ws[cell].fill = header_fill
            ws[cell].font = header_font
            ws[cell].alignment = center_align
            ws[cell].border = border
        
        # URL Status Data
        row = 10
        for url, status in data['url_status_summary'].items():
            ws[f'A{row}'] = url
            ws[f'B{row}'] = status['current_status']
            ws[f'C{row}'] = status['last_response_time']
            ws[f'D{row}'] = status['consecutive_failures']
            ws[f'E{row}'] = status['last_check']
            
            # Color code status
            status_cell = ws[f'B{row}']
            if status['current_status'] == 'UP':
                status_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            else:
                status_cell.fill = PatternFill(start_color="FFB6C6", end_color="FFB6C6", fill_type="solid")
            
            for cell in [f'A{row}', f'B{row}', f'C{row}', f'D{row}', f'E{row}']:
                ws[cell].border = border
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 25
        
        # Save file
        output_file = self.log_dir / "health_check_report.xlsx"
        wb.save(output_file)
        print(f"  [*] Excel report generated: {output_file}")
    
    def generate_html_report(self):
        """Generate HTML report"""
        if not self.report_file.exists():
            print("  ⚠ No report data available yet. Run a health check first.")
            return
        
        with open(self.report_file, 'r') as f:
            data = json.load(f)
        
        # Format timestamp for display
        def fmt(ts):
            try:
                return datetime.fromisoformat(ts).strftime('%Y-%m-%d %H:%M:%S')
            except Exception:
                return ts

        html_content = f"""<!DOCTYPE html>
<html lang='en'>
<head>
    <meta charset='UTF-8'>
    <meta name='viewport' content='width=device-width, initial-scale=1.0'>
    <title>IKEA Health Check Dashboard</title>
    <meta http-equiv='refresh' content='60'>
    <style>
        body {{ font-family: 'Segoe UI', Arial, sans-serif; background: #f4f6fa; margin: 0; }}
        .navbar {{ background: #0051BA; color: #fff; padding: 24px 0 16px 0; text-align: center; box-shadow: 0 2px 8px #0001; }}
        .navbar h1 {{ margin: 0; font-size: 2.2em; letter-spacing: 2px; }}
        .navbar .subtitle {{ font-size: 1.1em; color: #cce0ff; margin-top: 6px; }}
        .navbar .meta {{ font-size: 1em; color: #e0e0e0; margin-top: 10px; }}
        .container {{ width: 100vw; max-width: 100vw; margin: 0; background: none; border-radius: 0; box-shadow: none; overflow-x: hidden; }}
        .summary-cards {{ display: flex; flex-wrap: nowrap; overflow-x: auto; justify-content: center; gap: 16px; margin: 32px 0 24px 0; width: 100%; }}
        .card {{ background: #fff; border-radius: 10px; box-shadow: 0 2px 12px #0002; padding: 14px 12px; min-width: 120px; max-width: 140px; text-align: center; flex: 0 0 auto; }}
        .card-title {{ color: #888; font-size: 0.95em; margin-bottom: 6px; }}
        .card-value {{ font-size: 1.3em; font-weight: bold; color: #0051BA; }}
        .card-value.failures {{ color: #d32f2f; }}
        .card-value.healthy {{ color: #388e3c; }}
        .card-value.down {{ color: #d32f2f; }}
        .dashboard-table {{ width: 100vw; margin: 0 auto 40px auto; border-collapse: collapse; background: #fff; border-radius: 10px; overflow: hidden; box-shadow: 0 2px 12px #0001; }}
        .dashboard-table th, .dashboard-table td {{ padding: 12px 10px; text-align: left; }}
        .dashboard-table th {{ background: #0051BA; color: #fff; font-weight: 600; }}
        .dashboard-table tr:nth-child(even) {{ background: #f2f6fa; }}
        .dashboard-table td.status-up {{ color: #388e3c; font-weight: bold; }}
        .dashboard-table td.status-down {{ color: #d32f2f; font-weight: bold; }}
        .dashboard-table td.status-up::before {{ content: '🟢 '; }}
        .dashboard-table td.status-down::before {{ content: '🔴 '; }}
        .footer {{ text-align: center; color: #888; font-size: 0.98em; margin: 30px 0 10px 0; }}
        .copyright {{ text-align: center; color: #0051BA; font-size: 1em; margin: 10px 0 20px 0; font-weight: bold; }}
        @media (max-width: 900px) {{ .summary-cards {{ flex-wrap: nowrap; overflow-x: auto; }} .card {{ min-width: 110px; width: 110px; }} }}
        @media (max-width: 600px) {{ .dashboard-table th, .dashboard-table td {{ font-size: 0.95em; padding: 8px 4px; }} }}
    </style>
</head>
<body>
    <div class='navbar'>
        <h1>IKEA Health Check Dashboard</h1>
        <div class='subtitle'>Live Server Monitoring & Availability Report</div>
        <div class='meta'>
            Report generated: {fmt(data['timestamp'])} &nbsp;|&nbsp; Monitoring duration: {int(data['monitoring_duration_seconds'])} seconds
        </div>
    </div>
    <div class='container'>
    <div class='summary-cards'>
        <div class='card'>
            <div class='card-title'>Total URLs</div>
            <div class='card-value'>{data['monitored_urls']}</div>
        </div>
        <div class='card'>
            <div class='card-title'>Total Checks</div>
            <div class='card-value'>{data['total_checks']}</div>
        </div>
        <div class='card'>
            <div class='card-title'>Currently Healthy</div>
            <div class='card-value healthy'>{sum(1 for s in data['url_status_summary'].values() if s['current_status']=='UP')}</div>
        </div>
        <div class='card'>
            <div class='card-title'>Currently Down</div>
            <div class='card-value down'>{sum(1 for s in data['url_status_summary'].values() if s['current_status']=='DOWN')}</div>
        </div>
        <div class='card'>
            <div class='card-title'>Total Failures</div>
            <div class='card-value failures'>{data['total_failures']}</div>
        </div>
        <div class='card'>
            <div class='card-title'>Failure Rate</div>
            <div class='card-value'>{data['failure_rate']}</div>
        </div>
    </div>
    <table class='dashboard-table'>
        <thead>
            <tr>
                <th>#</th>
                <th>URL</th>
                <th>Status</th>
                <th>Status Code</th>
                <th>Response Time</th>
                <th>Consecutive Failures</th>
                <th>Last Check</th>
            </tr>
        </thead>
        <tbody>
"""
        for idx, (url, status) in enumerate(data['url_status_summary'].items(), 1):
            status_class = "status-up" if status['current_status'] == 'UP' else "status-down"
            html_content += f"<tr>"
            html_content += f"<td>{idx}</td>"
            html_content += f"<td>{url}</td>"
            html_content += f"<td class='{status_class}'>{status['current_status']}</td>"
            html_content += f"<td>{status['last_status_code']}</td>"
            html_content += f"<td>{status['last_response_time']}</td>"
            html_content += f"<td>{status['consecutive_failures']}</td>"
            html_content += f"<td>{fmt(status['last_check'])}</td>"
            html_content += f"</tr>"
        html_content += f"""
        </tbody>
    </table>
    </div>
    <div class='footer'>&nbsp;</div>
    <div class='copyright'>IKEA Health Check Monitoring Tool &copy; {datetime.now().year}</div>
</body>
</html>
"""
        output_file = self.log_dir / "index.html"
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(html_content)
        print(f"  [*] HTML report generated: {output_file}")
        # Trigger email alert after report generation
        try:
            import subprocess
            subprocess.Popen(['python', 'email_alert_win32.py'], cwd=str(Path(__file__).parent))
        except Exception as e:
            print(f"  [!] Failed to trigger email alert: {e}")


# ============================================================================
# MAIN MENU
# ============================================================================

def show_menu():
    """Display main menu"""
    print("\n" + "=" * 100)
    print("  IKEA HEALTH CHECK MONITORING TOOL")
    print("=" * 100)
    print("""
  1. Run Single Health Check Cycle
  2. Start Continuous Monitoring
  3. Generate Reports (Excel & HTML)
  4. View Recent Alerts
  5. Exit
  
  Enter your choice (1-5): """, end="")


def view_alerts(config: Dict):
    """Display recent alerts"""
    alert_file = Path(config['log_directory']) / config['alert_log_file']
    
    if not alert_file.exists():
        print("\n  ⚠ No alerts found yet.\n")
        return
    
    print("\n" + "=" * 100)
    print("  RECENT ALERTS")
    print("=" * 100)
    
    with open(alert_file, 'r') as f:
        lines = f.readlines()
        # Show last 20 alerts
        for line in lines[-20:]:
            print(f"  {line.rstrip()}")
    
    print("=" * 100 + "\n")


def main():
    """Main entry point"""
    print("\n")
    print("╔" + "=" * 98 + "╗")
    print("║" + " " * 20 + "IKEA SERVER HEALTH CHECK MONITORING TOOL" + " " * 37 + "║")
    print("║" + " " * 25 + "Unified Health Check & Reporting System" + " " * 34 + "║")
    print("╚" + "=" * 98 + "╝")
    
    # Load configuration
    config = ConfigLoader.load()
    
    while True:
        show_menu()
        choice = input().strip()
        
        if choice == '1':
            monitor = HealthCheckMonitor(config)
            if monitor.urls:
                monitor.run_single_check_cycle()
                monitor.print_summary()
            else:
                print("\n  ✗ ERROR: No URLs found in urls.txt\n")
        
        elif choice == '2':
            monitor = HealthCheckMonitor(config)
            if monitor.urls:
                monitor.start_continuous_monitoring()
            else:
                print("\n  ✗ ERROR: No URLs found in urls.txt\n")
        
        elif choice == '3':
            print("\n" + "=" * 100)
            print("  GENERATING REPORTS")
            print("=" * 100)
            generator = ReportGenerator(config)
            generator.generate_html_report()
            
            try:
                generator.generate_excel_report()
            except Exception as e:
                print(f"  ⚠ Excel report generation skipped: {str(e)}")
            
            print("=" * 100 + "\n")
        
        elif choice == '4':
            view_alerts(config)
        
        elif choice == '5':
            print("\n  Goodbye!\n")
            break
        
        else:
            print("\n  ✗ Invalid choice. Please enter 1-5.\n")


if __name__ == "__main__":
    # Handle command-line arguments
    if len(sys.argv) > 1:
        config = ConfigLoader.load()
        
        if sys.argv[1] == '--once':
            monitor = HealthCheckMonitor(config)
            if monitor.urls:
                monitor.run_single_check_cycle()
                monitor.print_summary()
        
        elif sys.argv[1] == '--continuous':
            monitor = HealthCheckMonitor(config)
            if monitor.urls:
                monitor.start_continuous_monitoring()
        
        elif sys.argv[1] == '--report':
            generator = ReportGenerator(config)
            generator.generate_html_report()
            try:
                generator.generate_excel_report()
            except:
                pass
        
        else:
            print("Usage:")
            print("  python run.py           - Interactive menu")
            print("  python run.py --once    - Run single check cycle")
            print("  python run.py --continuous - Continuous monitoring")
            print("  python run.py --report  - Generate reports")
    else:
        main()
